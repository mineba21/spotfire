"""
config/excel_utils.py

Excel (.xlsx) 다운로드 응답 생성 유틸리티.

메모리 효율 설계:
  1) openpyxl write_only 모드  → cell append 시점에 임시 XML 로 streaming write
  2) queryset.iterator()        → DB → Python 변환을 chunk 단위로 처리
  3) NamedTemporaryFile         → 완성된 .xlsx 를 디스크에 먼저 기록
  4) StreamingHttpResponse      → 디스크에서 chunk 단위로 클라이언트에 전송
                                   → HttpResponse 본문에 .xlsx 전체를 적재하지 않음

  이전 구현(wb.save(http_response)) 은 마지막 단계에서 .xlsx 바이트가
  HttpResponse 내부 버퍼에 누적되어, 동시 다운로드 시 서버 메모리가
  사용자 수 × 파일 크기 만큼 증가하는 문제가 있었다.
"""
import os
import tempfile
from urllib.parse import quote

from django.http import StreamingHttpResponse
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill


_HEADER_FONT  = Font(bold=True, color="FFFFFFFF")
_HEADER_FILL  = PatternFill("solid", fgColor="FF6366F1")
_HEADER_ALIGN = Alignment(horizontal="center")

_XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# 디스크 → 클라이언트 전송 시 read chunk 크기 (8KB 는 Django FileResponse 기본값과 동일)
_STREAM_CHUNK = 8192


def _stream_and_cleanup(path, chunk_size=_STREAM_CHUNK):
    """
    임시 파일을 chunk 단위로 yield 한 뒤, 응답 완료 시 파일을 삭제하는 generator.
    """
    try:
        with open(path, "rb") as f:
            while True:
                chunk = f.read(chunk_size)
                if not chunk:
                    break
                yield chunk
    finally:
        try:
            os.unlink(path)
        except OSError:
            # 이미 삭제됐거나 권한 문제 — 무시
            pass


def xlsx_response(rows, columns, sheet_name="Sheet1", filename="export.xlsx",
                  transform=None):
    """
    iterable 한 rows 를 .xlsx 로 변환해 StreamingHttpResponse 로 반환한다.

    파라미터
        rows       : dict 의 iterable. queryset.iterator() 도 OK.
        columns    : 출력할 컬럼 키 순서 (list of str).
        sheet_name : 워크시트 이름.
        filename   : 다운로드 파일명 (한글 OK — RFC 5987 인코딩 적용).
        transform  : (row_dict) -> dict 함수. None 이면 원본 사용.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name)

    # ── 헤더 행 (스타일 적용) ──────────────────────────────────
    header_cells = []
    for col in columns:
        cell = WriteOnlyCell(ws, value=col)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        header_cells.append(cell)
    ws.append(header_cells)

    # ── 데이터 행 ──────────────────────────────────────────────
    if transform is None:
        for row in rows:
            ws.append([row.get(c, "") for c in columns])
    else:
        for row in rows:
            t = transform(row)
            ws.append([t.get(c, "") for c in columns])

    # ── 임시 파일에 저장 → 디스크에서 stream ───────────────────
    # NamedTemporaryFile(delete=False) 로 핸들을 닫고 path 만 사용한다.
    # openpyxl 이 파일을 다시 열어서 zip 으로 직접 write 하므로 메모리 압박이 없다.
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()

    try:
        wb.save(tmp_path)
    except Exception:
        # save 실패 시 임시 파일 정리 후 예외 재발생
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise

    response = StreamingHttpResponse(
        _stream_and_cleanup(tmp_path),
        content_type=_XLSX_CONTENT_TYPE,
    )
    response["Content-Disposition"] = (
        f"attachment; filename*=UTF-8''{quote(filename)}"
    )
    # 일부 reverse-proxy 가 streaming response 를 buffering 하는 것을 막는다
    response["X-Accel-Buffering"] = "no"
    return response
